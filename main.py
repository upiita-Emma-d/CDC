import tkinter as tk
from tkinter import filedialog
import shapefile
from pyproj import Proj
import pandas as pd
from math import atan2, degrees, sqrt
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

fuente_general = Font(name='Eras Medium ITC', italic=False,size=11)

def calculate_polygon_area(coordinates):
    n = len(coordinates)
    area = 0.0
    for i in range(n):
        j = (i + 1) % n
        area += coordinates[i][0] * coordinates[j][1]
        area -= coordinates[j][0] * coordinates[i][1]
    area = abs(area) / 2.0
    return area

# Funciones de cálculo
def calculate_distance(point1, point2):
    return sqrt((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)

def calculate_bearing(point1, point2):
    angle_rad = atan2(point2[1] - point1[1], point2[0] - point1[0])
    angle_deg = degrees(angle_rad)
    bearing = (angle_deg + 360) % 360
    return bearing

def calculate_rumbo(bearing):
    """
    Convert a bearing angle in degrees to the traditional surveying notation.

    :param bearing: Bearing in degrees from the North line clockwise.
    :return: Bearing in the surveying format.
    """
    if 0 <= bearing < 90:
        return f"N {90-bearing:.5f}° E"
    elif 90 <= bearing < 180:
        return f"N {- 90 + bearing:.5f}° W"
    elif 180 <= bearing < 270:
        return f"S {270 - bearing:.5f}° W"
    elif 270 <= bearing < 360:
        return f"S {-270 + bearing:.5f}° E"
    else:
        raise ValueError("Bearing must be between 0 and 360 degrees.")


# Adjusted function for saving data to Excel to match the desired format
def save_to_excel(records, excel_output_path):
    wb = Workbook()
    ws = wb.active
    # Add the title "CUADRO DE CONSTRUCCION" at the top
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "CUADRO DE CONSTRUCCION"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Eras Medium ITC', size=16, bold=False, italic=False)

    # Add LADO
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    title_cell = ws.cell(row=2, column=1)
    title_cell.value = "LADO"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Eras Medium ITC', size=11, bold=False, italic=False)

    # Add EST
    # ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
    title_cell = ws.cell(row=3, column=1)
    title_cell.value = "EST"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Eras Medium ITC', size=11, bold=False, italic=False)

    # Add PV
    # ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
    title_cell = ws.cell(row=3, column=2)
    title_cell.value = "PV"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Eras Medium ITC',size=11, bold=False, italic=False)


    # Add RUMBO
    ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)
    title_cell = ws.cell(row=2, column=3)
    title_cell.value = "RUMBO"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Eras Medium ITC', size=11, bold=False, italic=False)

    # Add DISTANCIA
    ws.merge_cells(start_row=2, start_column=4, end_row=3, end_column=4)
    title_cell = ws.cell(row=2, column=4)
    title_cell.value = "DISTANCIA"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Eras Medium ITC', size=11, bold=False, italic=False)

    # Add V t
    # ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)
    title_cell = ws.cell(row=2, column=5)
    title_cell.value = "V"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Eras Medium ITC', size=11, bold=True, italic=False)

    # Add V t
    ws.merge_cells(start_row=2, start_column=5, end_row=3, end_column=5)
    title_cell = ws.cell(row=2, column=5)
    title_cell.value = "V"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Eras Medium ITC', size=11, bold=True, italic=False)

    # Add COORDENADAS
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)
    title_cell = ws.cell(row=2, column=6)
    title_cell.value = "COORDENADAS"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Eras Medium ITC', size=11, bold=True, italic=False)

    # Add X title
    # ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
    title_cell = ws.cell(row=3, column=6)
    title_cell.value = "X"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Eras Medium ITC', size=11, bold=True, italic=False)


    # Add Y title
    # ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
    title_cell = ws.cell(row=3, column=7)
    title_cell.value = "Y"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(name='Eras Medium ITC', size=11, bold=True, italic=False)


    # Add records to the Excel sheet
    for record in records:
        formatted_lon = "{:,.4f}".format(record["X"])
        formatted_lat = "{:,.4f}".format(record["Y"])
        if record["DISTANCIA"] != "":
            try:
                # Intenta convertir la distancia a un número flotante
                distancia_num = float(record["DISTANCIA"])
                # Si la conversión es exitosa, formatea el número
                formatted_dist = "{:,.3f}".format(distancia_num)
            except ValueError:
                # Si ocurre un ValueError durante la conversión, se maneja el error
                # Aquí puedes decidir qué hacer si no es un número, por ejemplo, dejarlo como cadena vacía
                formatted_dist = ""
        else:
            # Si es una cadena vacía, simplemente sigue adelante sin formatear
            formatted_dist = ""
        row = [
            record["EST"],
            record["PV"],
            record["RUMBO"],
            formatted_dist,
            record["V"],
            formatted_lon,
            formatted_lat,
        ]
        ws.append(row)

    # Calculate the area and add it at the bottom
    coordinates = [(record["X"], record["Y"]) for record in records]
    coordinates.append((records[0]["X"], records[0]["Y"]))  # Close the polygon by adding the first point at the end
    area = round(calculate_polygon_area(coordinates), 3)
    ws.merge_cells(start_row = (ws.max_row + 1) , start_column=ws.min_column, end_row= (ws.max_row + 1), end_column=ws.max_column)  
    superfaces_total = ws.cell(row=ws.max_row, column=ws.min_column)
    superfaces_total.value = f"SUPERFICIE {area} m2"
    superfaces_total.alignment = Alignment(horizontal='center', vertical='center')
    superfaces_total.font = Font(name='Eras Medium ITC', size=14, bold=False)
    altura_deseada = 30 
    ws.row_dimensions[ws.max_row].height = altura_deseada
    # Apply styles to all cells

    # Obtener el rango de filas y columnas
    min_row, max_row = 1, ws.max_row
    min_col, max_col = 1, 7
    borde_grueso = Side(style='thick')  # Borde más grueso para el contorno
    border_med = Side(style='medium')
    border_general = Side(style='thin')
    # Aplicar bordes gruesos solo al contorno externo
    for row in ws.iter_rows(min_row=min_row, max_col=max_col, min_col=min_col, max_row=max_row):
        for cell in row:
            # Aplicar alineación a todas las celdas
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Inicializar con bordes vacíos
            borders =  Border(
                left = border_general,
                right = border_general,
                top = border_general,
                bottom = border_general,
            )
            
            # Verificar si la celda la de abajo del titulo
            if cell.row == min_row + 1:
                    borders.top = border_med
            
            # Verificar si la celda es donde empiezan los datos
            if cell.row == min_row + 3:
                    borders.top = border_med        
                    
            # Verificar si la celda está en el borde izquierdo
            if cell.column == min_col:
                borders.left = borde_grueso
                
            # Verificar si la celda está en el borde derecho
            if cell.column == max_col:
                borders.right = borde_grueso
                
            # Verificar si la celda está en el borde superior
            if cell.row == min_row:
                borders.top = borde_grueso
            
            # Verificar si la celda está en el borde inferior
            if cell.row == max_row:
                borders.bottom = borde_grueso
                borders.top = border_med
            
            # Aplicar los bordes configurados a la celda
            cell.border = borders


    # Save the Excel file
    for row in ws.iter_rows():
        for cell in row[3:]:
            cell.font = fuente_general
    # Adjust the column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[1:-2])
        column_aux = column_cells[0].column
        letra = get_column_letter(column_aux)
        if letra in ("F", "G"):
            ws.column_dimensions[letra].width = 16
        elif letra in ("A", "B"):
            ws.column_dimensions[letra].width = 5
        # elif letra in ("D"):
        #     ws.column_dimensions[letra].width = 14
        else:
            ws.column_dimensions[letra].width = length + 2

            
    wb.save(excel_output_path)

# Función principal para procesar el shapefile
def process_shapefile(shapefile_path):
    sf = shapefile.Reader(shapefile_path)
    all_points = [shape.points[0] for shape in sf.shapes()]

    records = []
    record_init = {
        "EST": "",
        "PV": "",
        "RUMBO": "",
        "DISTANCIA": "",
        "V": 1,
        "X": all_points[0][0] if all_points[0] is not None else "",
        "Y": all_points[0][1] if all_points[0] is not None else "",
    }
    records.append(record_init)
    for i in range(len(all_points) - 1):
        point1, point2 = all_points[i], all_points[i + 1]
        distance = calculate_distance(point1, point2)
        bearing = calculate_bearing(point1, point2)
        rumbo = calculate_rumbo(bearing)
        lat1, lon1 = point1[1], point1[0]
        record = {
            "EST": i + 1, # or whatever value is appropriate for 'EST'
            "PV": i + 2, # or whatever value is appropriate for 'PV'
            "RUMBO": rumbo,
            "DISTANCIA": distance,
            "V": i + 2, # or the appropriate value for 'V'
            "X": lon1,
            "Y": lat1
        }
        records.append(record)
    point1, point2 =  all_points[-1],all_points[0]
    distance = calculate_distance(point1, point2)
    bearing = calculate_bearing(point1, point2)
    rumbo = calculate_rumbo(bearing)
    lat1, lon1 = point1[1], point1[0]
    record_end = {
        "EST": len(all_points),
        "PV": 1,
        "RUMBO": rumbo,
        "DISTANCIA": distance,
        "V": 1,
        "X": lon1,
        "Y": lat1
    }
    records.append(record_end)
    return records

# Función para seleccionar el shapefile usando una interfaz gráfica
def select_shapefile():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("Shapefiles", "*.shp")])
    root.destroy()  # Destroy the root window after selecting
    return file_path

# Flujo principal del programa
if __name__ == "__main__":
    shapefile_path = select_shapefile()
    if shapefile_path:
        print(f"Archivo seleccionado: {shapefile_path}")
        records = process_shapefile(shapefile_path)
        excel_output_path = 'salida.xlsx'
        save_to_excel(records, excel_output_path)
    else:
        print("No se seleccionó ningún archivo.")
